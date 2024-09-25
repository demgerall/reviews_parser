# pyuic6 C:/Users/demge/mainWindow.ui -o C:/Users/demge/PycharmProjects/TalkingBot/designMain.py

import datetime
import json
import time
import re
import os
import sys
import logging

from threading import Thread

from PyQt6 import QtWidgets

from selenium import webdriver as wd
from selenium.common import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    PatternFill, Border, Side,
    Alignment, Font
)
from selenium.webdriver.support.wait import WebDriverWait

import designMain


def set_styles_to_sheet(sheet: Worksheet, num: int) -> None:
    sheet.auto_filter.ref = f"A1:G{num - 1}"

    colors = ["DE0F10", "F48A11", "F49407", "ABBF1B", "015423"]
    cols = ["A", "B", "C", "D", "E", "F", "G"]

    for col in cols:
        for row in range(1, num):
            sheet[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            sheet[f"{col}{row}"].font = Font(name="Calibri", size=10)
            sheet[f"{col}{row}"].border = Border(left=Side(border_style="thin", color='000000'),
                                                 right=Side(border_style="thin", color='000000'),
                                                 top=Side(border_style="thin", color='000000'),
                                                 bottom=Side(border_style="thin", color='000000'), )
            if row == 1:
                sheet[f"{col}{row}"].alignment = Alignment(horizontal='center')
                sheet[f"{col}{row}"].font = Font(name="Calibri", size=12, bold=True)
                sheet[f"{col}{row}"].fill = PatternFill(patternType='solid', fgColor="FABF8F")
                sheet[f"{col}{row}"].border = Border(left=Side(border_style="medium", color='000000'),
                                                     right=Side(border_style="medium", color='000000'),
                                                     top=Side(border_style="medium", color='000000'),
                                                     bottom=Side(border_style="medium", color='000000'), )
            if col == "A" and row != 1:
                match sheet.cell(row=row, column=1).value:
                    case "+":
                        sheet[f"{col}{row}"].fill = PatternFill(patternType='solid', fgColor="ABBF1B")
                    case "-":
                        sheet[f"{col}{row}"].fill = PatternFill(patternType='solid', fgColor="DE0F10")
            if col == "D" and row != 1:
                sheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical="top")
                sheet[f"{col}{row}"].font = Font(color=colors[int(sheet.cell(row=row, column=4).value) - 1])
            if sheet.cell(row=row, column=cols.index(col) + 1).value == "None":
                sheet[f"{col}{row}"].font = Font(name="Calibri", size=12, bold=True, color="A94123")

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 150
    sheet.column_dimensions["F"].width = 20
    sheet.column_dimensions["G"].width = 100


def check_exists(el: WebElement, path: str) -> bool:
    try:
        el.find_element(By.CSS_SELECTOR, path)
        return True
    except NoSuchElementException:
        return False


class App(QtWidgets.QMainWindow, designMain.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.base_save_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        self.config = self.load_config()

        logging.basicConfig(level=logging.DEBUG, filename="logs.log", format="%(levelname)s (%(asctime)s): %(message)s (Line: %(lineno)d) [%(filename)s]", datefmt="%d/%m/%Y %I:%M:%S", encoding='UTF-8', filemode="a")

        self.save_textEdit.setPlaceholderText(
            f"Значение по умолчанию: {self.base_save_path}")
        self.filename_textEdit.setPlaceholderText(
            f"Значение по умолчанию: Отзывы <дата-время>.xlsx")

        self.start_button.clicked.connect(self.start)

    def load_config(self) -> object:
        try:
            with open('config.json', 'r') as f:
                config = json.load(f)
                self.status_label.setText("--Загрузка конфига прошла успешно--")
                return config
        except Exception as _ex:
            self.status_label.setText("--Загрузка конфига не прошла успешно--")
            self.error_label.setText("Возникла ошибка. Проверьте файл logs.log")
            logging.exception(_ex)
        finally:
            f.close()

    def create_excel_book(self) -> Workbook:
        try:
            book = openpyxl.Workbook()
            book.remove(book.active)
            return book
        except Exception as _ex:
            self.error_label.setText("Возникла ошибка. Проверьте файл logs.log")
            logging.exception(_ex)

    def save_excel_book(self, book: Workbook, path: str, excel_file_name: str) -> None:
        try:
            book.save(os.path.join(path, excel_file_name))
            self.status_label.setText(f"--Данные сохранены в {excel_file_name} по пути: {path}--")
        except Exception as _ex:
            self.error_label.setText("Возникла ошибка. Проверьте файл logs.log")
            logging.exception(_ex)
            self.status_label.setText(f"--Данные не удалось сохранить--")

    def start(self):
        self.error_label.setText("")

        with open("logs.log", "w", encoding='UTF-8') as f:
            f.write("")
        f.close()

        thread = Thread(target=self.start_parsing, daemon=True)
        thread.start()

    def start_parsing(self):
        book = self.create_excel_book()

        if self.filename_textEdit.text() == "":
            excel_file_name = f"Отзывы {datetime.datetime.now().strftime("%d-%b-%Y %H;%M;%S")}.xlsx"
        else:
            excel_file_name = self.filename_textEdit.text() + ".xlsx"

        if self.save_textEdit.text() == "":
            path = self.base_save_path
        else:
            path = self.save_textEdit.text()

        threads = []
        if self.gis_checkBox.isChecked():
            threads.append(Thread(target=self.search_on_site, args=(book, "2GIS",), daemon=True))
        if self.yandex_checkBox.isChecked():
            threads.append(Thread(target=self.search_on_site, args=(book, "Yandex",), daemon=True))
        if self.google_checkBox.isChecked():
            threads.append(Thread(target=self.search_on_site, args=(book, "Google",), daemon=True))

        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()

        self.save_excel_book(book=book, path=path, excel_file_name=excel_file_name)

        self.filename_textEdit.clear()
        self.save_textEdit.clear()

    def search_on_site(self, book: Workbook, site: str) -> None:
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")

            sheet = book.create_sheet(site)

            sheet.cell(row=1, column=1).value = "Ответ"
            sheet.cell(row=1, column=2).value = "Никнейм"
            sheet.cell(row=1, column=3).value = "Дата"
            sheet.cell(row=1, column=4).value = "Оценка"
            sheet.cell(row=1, column=5).value = "Текст отзыва"
            sheet.cell(row=1, column=6).value = "Дата ответа"
            sheet.cell(row=1, column=7).value = "Текст ответа"
            num = 2

            driver = wd.Chrome()
            driver.maximize_window()

            self.status_label.setText(f"--Поиск отзывов с {site} начался--")
            num = self.get_reviews_elements(site=site, num=num, sheet=sheet, driver=driver)

            set_styles_to_sheet(sheet=sheet, num=num)

        except Exception as _ex:
            self.error_label.setText("Возникла ошибка. Проверьте файл logs.log")
            logging.exception(_ex)

    def get_reviews_elements(self, site: str, num: int, sheet: Worksheet, driver: WebDriver) -> int:
        try:
            driver.get(url=self.config[site]["url"])

            action = ActionChains(driver)

            time.sleep(5)
            action.move_to_element(
                driver.find_element(By.CSS_SELECTOR, self.config[site]["clicked_element_css_selector"])).click()

            element = driver.find_element(By.CSS_SELECTOR, self.config[site]["scrolled_element_css_selector"])

            try:
                count_reviews = int(
                    driver.find_element(By.CSS_SELECTOR, self.config[site]["count_reviews_css_selector"]).text)
            except Exception as _ex:
                count_reviews = int(
                    "".join(re.findall(r'\d+', str(driver.find_element(By.CSS_SELECTOR, self.config[site][
                        "count_reviews_css_selector"]).text))))

            timer_start = time.perf_counter()

            while True:
                if len(driver.find_elements(By.CSS_SELECTOR, self.config[site]["show_more_button"])) > 0:
                    for el_to_open in driver.find_elements(By.CSS_SELECTOR, self.config[site]["show_more_button"]):
                        driver.execute_script("arguments[0].scrollIntoView(true);", el_to_open)
                        # el_to_open.click()
                        if site == "Google":
                            element = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, self.config[site]["show_more_button"]))
                            )
                            element.click()
                        else:
                            driver.execute_script("arguments[0].click();", el_to_open)
                if abs(len(driver.find_elements(By.CSS_SELECTOR, self.config[site][
                    "searched_card_css_selector"])) - count_reviews) > 3 and time.perf_counter() - timer_start < 60:
                    self.status_label.setText(
                        f"Прочитано {len(driver.find_elements(By.CSS_SELECTOR, self.config[site]["searched_card_css_selector"]))} отзывов из {count_reviews} с {site}...")
                    action.move_to_element(element).send_keys(Keys.END).perform()
                    time.sleep(0.5)
                else:
                    self.status_label.setText(
                        f"--Всего прочитано {len(driver.find_elements(By.CSS_SELECTOR, self.config[site]["searched_card_css_selector"]))} отзывов из {count_reviews} с {site}--")
                    self.status_label.setText(f"--Поиск отзывов с {site} закончился--")
                    num = self.get_reviews_data(driver=driver, site=site, num=num, sheet=sheet,
                                                html_els=driver.find_elements(By.CSS_SELECTOR,
                                                                              self.config[site][
                                                                                  "searched_card_css_selector"]))
                    return num

        except Exception as _ex:
            self.error_label.setText("Возникла ошибка. Проверьте файл logs.log")
            logging.exception(_ex)
            pass
        finally:
            driver.close()

    def get_reviews_data(self, driver: WebDriver, site: str, num: int, sheet: Worksheet,
                         html_els: list[WebElement]) -> int:
        self.status_label.setText(f"--Обработка данных с {site}--")

        for el in html_els:
            if check_exists(el, self.config[site]["review_answer_css_selector"]):
                sheet.cell(row=num, column=1).value = "+"  # review answer
            else:
                sheet.cell(row=num, column=1).value = "-"  # review answer
            if check_exists(el, self.config[site]["review_name_css_selector"]):
                sheet.cell(row=num, column=2).value = driver.execute_script('return arguments[0].textContent;',
                                                                            el.find_element(By.CSS_SELECTOR,
                                                                                            self.config[site][
                                                                                                "review_name_css_selector"]))  # name
            else:
                sheet.cell(row=num, column=2).value = "None"
            if check_exists(el, self.config[site]["review_date_css_selector"]):
                sheet.cell(row=num, column=3).value = driver.execute_script('return arguments[0].textContent;',
                                                                            el.find_element(By.CSS_SELECTOR,
                                                                                            self.config[site][
                                                                                                "review_date_css_selector"])).replace(
                    ", отредактирован", "(отредактирован)")  # date
            else:
                sheet.cell(row=num, column=3).value = "None"
            if check_exists(el, self.config[site]["review_rate_css_selector"]):
                sheet.cell(row=num, column=4).value = len(
                    el.find_elements(By.CSS_SELECTOR, self.config[site]["review_rate_css_selector"]))  # rate
            else:
                sheet.cell(row=num, column=4).value = "None"
            if check_exists(el, self.config[site]["review_text_css_selector"]):
                sheet.cell(row=num, column=5).value = driver.execute_script('return arguments[0].textContent;',
                                                                            el.find_element(By.CSS_SELECTOR,
                                                                                            self.config[site][
                                                                                                "review_text_css_selector"]))  # text review
            else:
                sheet.cell(row=num, column=5).value = "None"
            if check_exists(el, self.config[site]["review_answer_date_css_selector"]):
                sheet.cell(row=num, column=6).value = driver.execute_script('return arguments[0].textContent;',
                                                                            el.find_element(By.CSS_SELECTOR,
                                                                                            self.config[site][
                                                                                                "review_answer_date_css_selector"]))  # answer date
            else:
                sheet.cell(row=num, column=6).value = "None"
            if check_exists(el, self.config[site]["review_answer_text_css_selector"]):
                sheet.cell(row=num, column=7).value = driver.execute_script('return arguments[0].textContent;',
                                                                            el.find_element(By.CSS_SELECTOR,
                                                                                            self.config[site][
                                                                                                "review_answer_text_css_selector"]))  # answer text
            else:
                sheet.cell(row=num, column=7).value = "None"

            num += 1

        self.status_label.setText(f"--Обработка данных с {site} завершена--")
        return num


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = App()
    window.show()
    app.exec()


if __name__ == '__main__':
    main()
